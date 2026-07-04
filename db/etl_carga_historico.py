"""
ETL — Carga dos dados históricos para o Supabase
Projeto: Sistema Financeiro E-commerce

Como usar:
1. pip install supabase python-dotenv openpyxl
2. Crie .env com SUPABASE_URL e SUPABASE_KEY
3. Coloque a planilha DRE_ECOMMERCE_OFICIAL.xlsx na mesma pasta
4. python etl_carga_historico.py
"""

import os
from openpyxl import load_workbook
from supabase import create_client, Client
from dotenv import load_dotenv
from datetime import datetime

# ── Configuração ──────────────────────────────────────────────────────────────
load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
PLANILHA     = "DRE_ECOMMERCE_OFICIAL.xlsx"

if not SUPABASE_URL or not SUPABASE_KEY:
    raise EnvironmentError("Configure SUPABASE_URL e SUPABASE_KEY no arquivo .env")

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ── Helpers ───────────────────────────────────────────────────────────────────
def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

def safe_float(val) -> float:
    try:
        if val is None: return 0.0
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def competencia_para_db(mes_str: str) -> str:
    """Converte 'Jan/25' → '2025-01'"""
    meses = {
        "Jan":"01","Fev":"02","Mar":"03","Abr":"04",
        "Mai":"05","Jun":"06","Jul":"07","Ago":"08",
        "Set":"09","Out":"10","Nov":"11","Dez":"12"
    }
    p   = mes_str.strip().split("/")
    mes = meses.get(p[0], "01")
    ano = f"20{p[1]}"
    return f"{ano}-{mes}"

# ── Buscar IDs cadastrados no banco ──────────────────────────────────────────
def buscar_canais() -> dict:
    res = supabase.table("canais").select("id, nome").execute()
    return {r["nome"]: r["id"] for r in res.data}

def buscar_funcionarios() -> dict:
    res = supabase.table("funcionarios").select("id, nome").execute()
    return {r["nome"]: r["id"] for r in res.data}

def buscar_categorias() -> dict:
    res = supabase.table("categorias_despesa").select("id, nome").execute()
    return {r["nome"]: r["id"] for r in res.data}

# ── Leitura do INPUT_PLATAFORMAS ──────────────────────────────────────────────
def ler_receitas(wb) -> list:
    ws       = wb["INPUT_PLATAFORMAS"]
    canais_db = buscar_canais()

    # Meses: linha 2, colunas D(4) a O(15)
    meses = []
    for col in range(4, 16):
        val = ws.cell(row=2, column=col).value
        if val:
            meses.append((col, str(val)))

    # Mapeamento fixo: canal → linha inicial dos campos
    # Cada canal tem 9 linhas: RB, Imp, CP, Fr, Ta, Dc, Af, Ou, Ads
    CANAIS_MAP = {
        "Mercado Livre": 3,
        "Shopee":        14,
        "Shein":         25,
        "TikTok":        36,
        "Bazar":         47,
        "Site":          58,
        "WhatsApp":      69,
    }
    # Offsets relativos à linha inicial do canal
    CAMPOS_OFFSET = {
        "receita_bruta":  0,
        "imposto":        1,
        "custo_produto":  2,
        "frete":          3,
        "tarifa":         4,
        "desconto_cupom": 5,
        "afiliado":       6,
        "outros_cmv":     7,
        "ads":            8,
    }

    registros = []
    for canal, row_inicio in CANAIS_MAP.items():
        canal_id = canais_db.get(canal)
        if not canal_id:
            log(f"  ⚠ Canal '{canal}' não encontrado no banco — pulando")
            continue

        for col, mes_str in meses:
            comp = competencia_para_db(mes_str)
            rec  = {"canal_id": canal_id, "competencia": comp}

            for campo, offset in CAMPOS_OFFSET.items():
                val = ws.cell(row=row_inicio + offset, column=col).value
                rec[campo] = safe_float(val)

            # Só inserir se houver receita
            if rec["receita_bruta"] > 0:
                registros.append(rec)

    return registros

# ── Leitura do INPUT_DESPESAS — Folha ─────────────────────────────────────────
def ler_folha(wb) -> list:
    ws       = wb["INPUT_DESPESAS"]
    funcs_db = buscar_funcionarios()

    # Meses: linha 2, colunas D(4) a O(15)
    meses = []
    for col in range(4, 16):
        val = ws.cell(row=2, column=col).value
        if val:
            meses.append((col, str(val)))

    # Mapeamento fixo das linhas por funcionário
    FUNC_MAP = {
        "Ben-Hur": {"salario": 4,  "bonus": 20, "fgts": None, "fgts_bonus": None, "inss": None},
        "Camila":  {"salario": 6,  "bonus": 19, "fgts": None, "fgts_bonus": None, "inss": None},
        "Jackson": {"salario": 8,  "bonus": 15, "fgts": 9,    "fgts_bonus": 16,   "inss": 10},
        "Laura":   {"salario": 11, "bonus": 17, "fgts": 12,   "fgts_bonus": 18,   "inss": 13},
    }

    registros = []
    for nome, linhas in FUNC_MAP.items():
        func_id = funcs_db.get(nome)
        if not func_id:
            log(f"  ⚠ Funcionário '{nome}' não encontrado no banco — pulando")
            continue

        for col, mes_str in meses:
            comp    = competencia_para_db(mes_str)
            sal     = safe_float(ws.cell(row=linhas["salario"], column=col).value)
            bonus   = safe_float(ws.cell(row=linhas["bonus"],   column=col).value) if linhas["bonus"] else 0.0
            fgts    = safe_float(ws.cell(row=linhas["fgts"],    column=col).value) if linhas["fgts"] else 0.0
            fgts_bn = safe_float(ws.cell(row=linhas["fgts_bonus"], column=col).value) if linhas["fgts_bonus"] else 0.0
            inss    = safe_float(ws.cell(row=linhas["inss"],    column=col).value) if linhas["inss"] else 0.0

            if sal + bonus > 0:
                registros.append({
                    "funcionario_id": func_id,
                    "competencia":    comp,
                    "salario_bruto":  sal,
                    "bonus":          bonus,
                    "fgts":           fgts + fgts_bn,
                    "inss_retido":    inss,
                })

    return registros

# ── Leitura do INPUT_DESPESAS — Operacional e Marketing ───────────────────────
def ler_despesas(wb) -> list:
    ws      = wb["INPUT_DESPESAS"]
    cats_db = buscar_categorias()

    meses = []
    for col in range(4, 16):
        val = ws.cell(row=2, column=col).value
        if val:
            meses.append((col, str(val)))

    # Mapeamento fixo: (descricao, categoria, tipo, linha)
    DESPESAS_MAP = [
        ("Aluguel",              "Administrativo",         "fixo",     23),
        ("Ferramentas",          "Administrativo",         "fixo",     24),
        ("Contador",             "Administrativo",         "fixo",     25),
        ("Embalagem",            "Administrativo",         "variavel", 26),
        ("Trilha do Ecommerce",  "Educação e Capacitação", "eventual", 27),
        ("Outros Operacionais",  "Administrativo",         "eventual", 28),
        ("Tráfego Pago Externo", "Marketing",              "variavel", 31),
        ("Outros Marketing",     "Marketing",              "variavel", 32),
    ]

    registros = []
    for (desc, categoria, tipo, linha) in DESPESAS_MAP:
        cat_id = cats_db.get(categoria)
        if not cat_id:
            log(f"  ⚠ Categoria '{categoria}' não encontrada — pulando")
            continue

        for col, mes_str in meses:
            comp = competencia_para_db(mes_str)
            val  = safe_float(ws.cell(row=linha, column=col).value)
            if val > 0:
                registros.append({
                    "categoria_id": cat_id,
                    "competencia":  comp,
                    "descricao":    desc,
                    "tipo":         tipo,
                    "valor":        val,
                })

    return registros

# ── Carga no Supabase ─────────────────────────────────────────────────────────
def upsert_receitas(registros: list):
    if not registros:
        log("  Nenhum registro de receita para carregar.")
        return
    supabase.table("receitas").upsert(
        registros, on_conflict="canal_id,competencia"
    ).execute()
    log(f"  ✓ {len(registros)} registros de receita carregados")

def upsert_folha(registros: list):
    if not registros:
        log("  Nenhum registro de folha para carregar.")
        return
    supabase.table("folha_pagamento").upsert(
        registros, on_conflict="funcionario_id,competencia"
    ).execute()
    log(f"  ✓ {len(registros)} registros de folha carregados")

def inserir_despesas(registros: list):
    if not registros:
        log("  Nenhuma despesa para carregar.")
        return
    # Deletar e reinserir por competência para evitar duplicatas
    competencias = list(set(r["competencia"] for r in registros))
    for comp in competencias:
        supabase.table("despesas").delete().eq("competencia", comp).execute()
    supabase.table("despesas").insert(registros).execute()
    log(f"  ✓ {len(registros)} despesas carregadas")

def recalcular_dre(competencias: list):
    for comp in sorted(set(competencias)):
        supabase.rpc("recalcular_dre", {"p_competencia": comp}).execute()
        log(f"  ✓ DRE consolidada recalculada: {comp}")

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    log("=== Iniciando carga histórica ===")
    log(f"Planilha: {PLANILHA}")

    wb = load_workbook(PLANILHA, data_only=True)

    log("\n[1/4] Lendo receitas por plataforma...")
    receitas = ler_receitas(wb)
    log(f"  {len(receitas)} registros encontrados")
    upsert_receitas(receitas)

    log("\n[2/4] Lendo folha de pagamento...")
    folha = ler_folha(wb)
    log(f"  {len(folha)} registros encontrados")
    upsert_folha(folha)

    log("\n[3/4] Lendo despesas operacionais...")
    despesas = ler_despesas(wb)
    log(f"  {len(despesas)} registros encontrados")
    inserir_despesas(despesas)

    log("\n[4/4] Recalculando DRE consolidada...")
    competencias = list(set(r["competencia"] for r in receitas))
    recalcular_dre(competencias)

    log("\n=== Carga concluída! ===")
    log(f"Meses processados: {', '.join(sorted(set(competencias)))}")

if __name__ == "__main__":
    main()